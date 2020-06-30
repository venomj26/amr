#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Apr 10 02:24:25 2020

@author: jha
"""


#%%
import csv
import mysql.connector as mariadb
from tabulate import tabulate
import pandas as pd

amrdb= mariadb.connect(
        host="localhost",
        user="root",
        passwd="Sukhoi@90",
        database ="myamr"
        )
cursor = amrdb.cursor(buffered=True) # else it fetches one row for everytime it is executed 
print("We are at line 31 we have connection, lets begin")

#%%


#get the data from the ast table to make the dataframe 
#%%
import pandas as pd
numrows= cursor.execute("Describe Escherichia_nov_astt")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
print(rows)
df_col_name= pd.DataFrame(rows)
list_col_name=list(df_col_name[0])

del list_col_name[0]
del list_col_name[0]
date_df=[]
for item in list_col_name:
    try:
        numrows= cursor.execute("SELECT " + item +" FROM Escherichia_isolateepi where isolate_epi_type ='clinical'")
        print("SELECT " + item +" FROM Escherichia_isolateepi")
        print("Selected %s rows" %numrows)
        print("Selected %s rows " %cursor.rowcount)
        rows =cursor.fetchall()#fetch all rows at once
        ast_df=pd.DataFrame(rows)
        ast_df.columns=[item]
        date_df.append(ast_df)
    
    except:
        print("no drug for you")
        continue

datadf=pd.concat(date_df, sort=False, axis=1)

#print(tabulate(rows, headers=['AST_phenotype'], tablefmt='psql'))
#%%



        
#get the data from the ast table to make the dataframe 
#%%
import pandas as pd
numrows= cursor.execute("SELECT Collection_year FROM Escherichia_isolateepi where isolate_epi_type ='clinical'")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
df=pd.DataFrame(rows)
df.columns=["Collection_year"]
datadf=datadf.join(df)
#%%    
    
    
    
    


#best way to read table from sql
#%%
import pandas as pd
numrows= cursor.execute("Describe Ecolinew")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
print(rows)
dfz= pd.DataFrame(rows)

listz=list(dfz[0])
numrows_k= cursor.execute("select * from Ecolinew where `#label` like '%PDT000198421.3%';")
print("Selected %s rows" %numrows_k)
print("Selected %s rows " %cursor.rowcount)
rows_k =cursor.fetchall()#fetch all rows at once
df_kleb= pd.DataFrame(rows_k)
df_kleb.columns= listz

#%%




#checking the duplicates in file needs the unique IDs to be removed 
#%%
df_kleb=df_kleb.drop(['ast_code', 'astcode', 'id', 'astcode2', 'astcode3', 'astcode4'], axis = 1)
#from sqlalchemy import create_engine
#cnx = create_engine('mysql+pymysql://root:Sukhoi@90@localhost/myamr')
#df_kleb.to_sql(name='checkingtableecoli', con=cnx, if_exists = 'replace', index=False)
#print("we are at line 186 in code")
#df_kleb=df_kleb.drop_duplicates(inplace=False)
df_kleb.to_excel("Escherichiacoli_PDT000198421*.3.xlsx")
#%%




#%%
#temp_df = pd.pivot_table(datadf, values ='imipenem', index =["Collection_year","isolate_epi_type"], 
 #                        columns =["isolate_epi_type"], aggfunc=np.sum) 

df_crosstab=pd.crosstab( datadf.Collection_year,datadf.isolate_epi_type, margins=True, margins_name="total samples")


#%%



#%%
import pandas as pd
numrows= cursor.execute("Describe Escherichia_nov_astt")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
print(rows)
df_col_name= pd.DataFrame(rows)
list_col_name=list(df_col_name[0])

del list_col_name[0]
del list_col_name[0]


#%%


#for getting all the antimicrobials
#%%
 #def crosstab_all(df_kleb,list_col_name):
appended_data=[]
for k in list_col_name:
    try:
        xdf=pd.crosstab(datadf["Collection_year"],datadf[k], margins=True, margins_name="Out of total samples")
        xdf["Antimicrobial"] = k
        print('xdf',xdf)
        print('') # for spacing
        appended_data.append(xdf)
       
             
    except:
        print("not found")
        continue
appended_data = pd.concat(appended_data, sort= False) 
#appended_data.loc[appended_data['Antimicrobial'].duplicated(), 'Antimicrobial'] = np.nan

#%%


#for only those antimicrobials which were tested for more than one epitypes
#%%
 #def crosstab_all(df_kleb,list_col_name):
 appended_data=[]
for k in list_col_name:
    try:
        xdf=pd.crosstab(datadf["Collection_year"],datadf[k], margins=True, margins_name="Out of total samples")
        xdf["Antimicrobial"] = k
        print('xdf',xdf)
        print('') # for spacing
        if len(xdf.index) >2:
            appended_data.append(xdf)
        else:
            print("only one epitype tested")
             
    except:
        print("not found")
        continue
appended_data= pd.concat(appended_data, sort= False) 
appended_data=appended_data.drop("missing")
appended_data=appended_data.drop("")
#appended_data_P=appended_data_P.drop("DD", axis=1)

#%%


#%%

df_family = pd.read_excel('/Users/jha/Documents/spring2020/ecoli_antimicrobial_class.xlsx', sheet_name='Sheet1')
df_family_mod= df_family
#df_family=df_family.drop(["Intrinsicly_resistant","Antimicrobial_class","Antimicrobial_mod"], axis=1)
list_family=list(df_family.itertuples(index=False, name=None))
#list_family=list(df_family)
family_dict=dict(list_family)

appended_data['Antimicrobial_class']=appended_data['Antimicrobial'].map(family_dict)
appended_data = appended_data.reset_index()

appended_data= (appended_data.groupby(['Antimicrobial_class', 'Antimicrobial','Collection_year']).sum())
appended_data = appended_data.reset_index()

appended_data.loc[appended_data['Antimicrobial'].duplicated(), 'Antimicrobial'] = np.nan
appended_data.loc[appended_data['Antimicrobial_class'].duplicated(), 'Antimicrobial_class'] = np.nan

appended_data = appended_data.set_index(['Antimicrobial_class', 'Antimicrobial'])


cols=["Collection_year","S","R","I","Out of total samples"]
appended_data =appended_data[cols]
 

#%%



#%%
import pandas as pd
from openpyxl import load_workbook
path=  "/Users/jha/escherichia_isolation_antimicrobial_crosstab.xlsx"


book = load_workbook(path)
writer = pd.ExcelWriter(path, engine = 'openpyxl')
writer.book = book


appended_data.to_excel(writer, sheet_name="temporal_clinical_multiyears")
writer.save()
writer.close()
    
#%%

#for only those antimicrobials which were tested for more than one epitypes
#%%
 #def crosstab_all(df_kleb,list_col_name):
 appended_data_P=[]
for k in list_col_name:
    try:
        xdf=pd.crosstab(datadf["Collection_year"],datadf[k], margins=True, margins_name="Out of total samples", normalize="index").round(4)*100
        xdf["Antimicrobial"] = k
        print('xdf',xdf)
        print('') # for spacing
        if len(xdf.index) >2:
            appended_data_P.append(xdf)
        else:
            print("only one epitype tested")
             
    except:
        print("not found")
        continue
appended_data_P= pd.concat(appended_data_P , sort= False) 
appended_data_P=appended_data_P.drop("missing")
appended_data_P=appended_data_P.drop("")
appended_data_P=appended_data_P.drop("DD", axis=1)

#%%

#to merge the appended tables to get frequency and percentage in the same table
#%%

appended_data_P=appended_data_P.rename(columns = {"S": "%Susceptible", 
                                  "R":"%Resistant", 
                                  "I": "%Intermediate"})
appended_data=appended_data.reset_index()
appended_data=appended_data.reset_index()
appended_data_P=appended_data_P.reset_index()
appended_data_P=appended_data_P.reset_index()
appended_dict = dict(zip(appended_data.index, appended_data.S))
appended_data_P["Susceptible"]=appended_data_P['index'].map(appended_dict)

appended_dict = dict(zip(appended_data.index, appended_data.R))
appended_data_P["Resistant"]=appended_data_P['index'].map(appended_dict)

appended_dict = dict(zip(appended_data.index, appended_data.I))
appended_data_P["Intermediate"]=appended_data_P['index'].map(appended_dict)

appended_data_P=appended_data_P.drop("index", axis=1)
appended_data=appended_data.drop("index", axis=1)
#appended_data = appended_data.set_index(['Collection_year'])

#%%


#%%
#
df_family = pd.read_excel('/Users/jha/Documents/spring2020/ecoli_antimicrobial_class.xlsx', sheet_name='Sheet1')
df_family_mod= df_family
#df_family=df_family.drop(["Intrinsicly_resistant","Antimicrobial_class","Antimicrobial_mod"], axis=1)
list_family=list(df_family.itertuples(index=False, name=None))
#list_family=list(df_family)
family_dict=dict(list_family)

appended_data_P['Antimicrobial_class']=appended_data_P['Antimicrobial'].map(family_dict)

appended_data_P= (appended_data_P.groupby(['Antimicrobial_class', 'Antimicrobial','Collection_year']).sum())
appended_data_P = appended_data_P.reset_index()
#%%

#%%
appended_data_P.loc[appended_data_P['Antimicrobial'].duplicated(), 'Antimicrobial'] = np.nan
appended_data_P.loc[appended_data_P['Antimicrobial_class'].duplicated(), 'Antimicrobial_class'] = np.nan

#%%

#%%


appended_data_P = appended_data_P.set_index(['Antimicrobial_class', 'Antimicrobial'])
#%%


#%%
cols=["Collection_year","%Susceptible","Susceptible","%Resistant","Resistant","%Intermediate","Intermediate"]
appended_data_P =appended_data_P[cols]
 

#%%


#get the data from the ast table to make the dataframe 
#%%
import pandas as pd
numrows= cursor.execute("Describe Ecoli2020")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
print(rows)
dfz= pd.DataFrame(rows)
listz=list(dfz[0])

numrows= cursor.execute("select * from Ecoli2020 where `#label` = 'PDT000002365.3|SAMN02368173||Escherichia coli||ESC0165|Pathogen: clinical or host-associated'")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
df=pd.DataFrame(rows)
#df.columns=["bioproject-name","samplename"]
df.columns= listz

#%%    
    





#%%
import pandas as pd
from openpyxl import load_workbook
path=  "/Users/jha/Escherichiacoli2020.xlsx"


book = load_workbook(path)
writer = pd.ExcelWriter(path, engine = 'openpyxl')
writer.book = book


df.to_excel(writer, sheet_name="PDT000002365.3")
writer.save()
writer.close()
    
#%%
