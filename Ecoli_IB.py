#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue May 12 11:13:35 2020

@author: jha
"""

#used to write the IB excel file into sql 
#%%

#df_family = pd.read_excel('/Users/jha/Documents/spring2020/Klebsiella_Sneha_Jan_2020.xlsx', sheet_name='Sheet1')
import pandas as pd
df=pd.DataFrame.from_csv('/Users/jha/Documents/spring2020/data/ecoli_IB/Pathogen.tsv', sep='\t', header=0)

df=pd.DataFrame.from_csv('/Users/jha/Documents/spring2020/data/ecoli_IB/Ecoli_IB_AMR.txt', sep='\t', header=0)
df=df.reset_index()
#%%

#%%

df=df.replace(r'^\s*$', np.nan, regex=True)
from sqlalchemy import create_engine
cnx = create_engine('mysql+pymysql://root:Sukhoi@90@localhost/myamr')
df.to_sql(name='EcoliIB_amr', con=cnx, if_exists = 'replace', index=False)
print("we are at line 186 in code")





#%%

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun May 10 23:15:01 2020

@author: jha
"""

#!/usr/bin/env python3
# -*- coding: utf-8 -*-



#%%
import csv
import mysql.connector as mariadb
from tabulate import tabulate
import pandas as pd

amrdb= mariadb.connect(
        host="localhost",
        user="root",
        passwd="Sukhoi@90",
        database ="myamr"
        )
cursor = amrdb.cursor(buffered=True) # else it fetches one row for everytime it is executed 
print("We are at line 31 we have connection, lets begin")

#%%



#creating a table where all AST_phenotypes values are from USA
#%%
new_table= cursor.execute("CREATE TABLE IF NOT EXISTS Ecoli_IB_us AS SELECT * FROM Ecoli_IB  WHERE Location LIKE '%US%'  ")
print("Selected %s rows" %new_table)
print("Selected %s rows " %cursor.rowcount) 
#rows =cursor.fetchall()#fetch all rows at once
#print(tabulate(rows, headers=['strain','AST_phenotype', 'AMR_genotype'], tablefmt='psql'))

#%%




#get the data from the ast table to make the dataframe 
#%%
import pandas as pd
numrows= cursor.execute("SELECT Virulence_genotypes FROM EcoliIBisodatlocmerge")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
#print(tabulate(rows, headers=['AST_phenotype'], tablefmt='psql'))
df = pd.DataFrame(rows)
#%%


#splitting the AST_phenotype column values into separate columns of the table 
#%%
df = pd.DataFrame(rows)
#print ("the pandas dataframe ",df
       
df.columns=["AST_phenotypes"]
df.head()
df=df.replace(r'^\s*$','VALUE=DNE', regex=True)

df= df.fillna(0)
df=df.replace(0, 'VALUE=ND', regex=True)
#%%

#%%
df['ASTphenotypes']=df['AST_phenotypes']
df=df.set_index('ASTphenotypes')
split_df= df['AST_phenotypes'].str.split(',', expand= True)
#split_df.to_excel("split_df_ecoli_IB.xlsx") #used to write into excel file 
split_df.head()
split_dfT=split_df.T
#split_df = split_df.drop(split_df.columns[[23]], axis=1) 

#%%


#we are making the column names for the AST dataframes 
#%%
split_df_list=[]
split_df_list = split_df.values.tolist()
flat_list_sd = []
for sublist in split_df_list:
    for item in sublist:
        flat_list_sd.append(item)

#print ("ATline {0} in code split_df has been converted to a list{1}".format(lineno(),flat_list_sd))
split_flsd=[]
for element in flat_list_sd:
    if element != None:
        split_flsd.append(element.split('=', 1)[0]) 
#print ("ATline {0} in code split_df has been converted to a list".format(lineno()))
print("the list of all elements in split_df", split_flsd)
header_set=set(split_flsd)
header_list_AST=list(header_set)

print("at line 87 in code the list of header values are ")

df_AST=pd.DataFrame(header_list_AST)
df_AST.columns=["AST"]
df_AST=df_AST.T
df_AST=df_AST.reset_index()
df_AST=df_AST.T
df_AST=df_AST.reset_index()
df_AST=df_AST.drop(['index'], axis=1)
df_AST.columns=["AST"]


#%%




#%%
x=1

for tup in split_df.itertuples():
    try:
        print (tup)
        tuplist=list(tup)
        tup_list_df=pd.DataFrame(tuplist)
        colname=tup_list_df.iloc[0]
        head= pd.DataFrame([colname])
        head=head.reset_index()
        head= head.T
        head=head.reset_index()
        print("the head at A looks like /n",head)
        head=head.drop(columns=['index'])
        print("the head at B looks like/n",head)
        head=head.set_value(0,0, 'AST')
        print("the head at c looks like /n",head)
        head.columns=["header"]
        print("the head at D looks like /n",head)
        head_list=head['header'].tolist()
        print("the head_list at E looks like /n",head_list)
        tup_list_df=tup_list_df.drop([0])
        tup_list_df.columns=["AST"]
        tup_list_df= tup_list_df.AST.str.split('=', expand= True)
        tup_list_df.columns=head_list
        tup_list_df=tup_list_df.T
        tup_list_df=tup_list_df.reset_index()
        tup_list_df=tup_list_df.T
        tup_list_df.columns=["AST","AST_p"]
        #tup_list_df_list= list(tup_list_df)
        tup_list_df_list=list(tup_list_df.itertuples(index=False, name=None))
        tup_list_df_list_dict=dict(tup_list_df_list)
        #tup_list_df_list_dict=dict(tup_list_df_list)
        df_AST[x]=df_AST['AST'].map(tup_list_df_list_dict)
        x=x+1
        
    except:
        print("the header list has a problem")
        continue


#%%

#%%
import pandas as pd
from openpyxl import load_workbook
path=  "/Users/jha/split_df_ecili_IB.xlsx"


book = load_workbook(path)
writer = pd.ExcelWriter(path, engine = 'openpyxl')
writer.book = book


df_AST.to_excel(writer, sheet_name="df_ast")
writer.save()
writer.close()
    
#%%

#%%
df_ASTT=df_AST.T
df_ASTT.columns = df_ASTT.iloc[0] #grab the first row for the header

df_ASTT=df_ASTT.drop(['AST'])# drops the first row which is a header row from dataframe but this also changes the index of the dataframe
df_ASTT=df_ASTT.reset_index()# the change of index in the previous command causes error in the concatenation of the other dataframe
df_ASTT= df_ASTT.drop(columns=['index'])# so reset is used to create a new index and the old index is deleted 
#df_ASTT = pd.concat([df,df_ASTT], axis=1, join_axes=[df.index])



#%%

#%%
import pandas as pd
from openpyxl import load_workbook
path=  "/Users/jha/split_df_ecoli_IB_check.xlsx"


book = load_workbook(path)
writer = pd.ExcelWriter(path, engine = 'openpyxl')
writer.book = book


df_AST.to_excel(writer, sheet_name="df_astt")
writer.save()
writer.close()
    
#%%















#This removes all drugs which have only none values in all rows and columns 
#%%

#df_ASTT = df_ASTT.replace('ND', np.nan)
#df_ASTT = df_ASTT.fillna('0')
#df_ASTT = df_ASTT.replace('0', np.nan)
#df_ASTT = df_ASTT.set_index('AST_phenotypes')
#df_ASTT = df_ASTT.dropna(axis = 0, how = 'all')
#df_ASTT = df_ASTT.dropna(axis = 1, how = 'all')
#df_ASTT = df_ASTT.reset_index()

#%%



#run this code only once to create the astt table
#%%
df_ASTT= df_ASTT.rename(columns={"AST":"Genotype"})
df_ASTT=df_ASTT.replace(r'^\s*$', np.nan, regex=True)

from sqlalchemy import create_engine
cnx = create_engine('mysql+pymysql://root:Sukhoi@90@localhost/myamr')
df_ASTT.to_sql(name='Ecoli_IB_AMR', con=cnx, if_exists = 'replace', index=False)
print("we are at line 186 in code")

#%%

#%%
sql="ALTER TABLE `Ecoli_IB` ADD COLUMN `amrid` INT AUTO_INCREMENT UNIQUE FIRST;"
sql="CREATE TABLE IF NOT EXISTS Ecoli_IB_merged AS  SELECT * FROM EcoliIB_amr   INNER JOIN Ecoli_IB_us ON  EcoliIB_amr.amr_id = Ecoli_IB_us.amrid;"
SQL="ALTER TABLE EcoliIB_amr CHANGE COLUMN `AMR_genotypes` `GENOTYPE` longtext;"
SQL="CREATE TABLE IF NOT EXISTS Ecoli_IB_merged AS  SELECT * FROM EcoliIB_amr   INNER JOIN Ecoli_IB_us ON  EcoliIB_amr.amr_id = Ecoli_IB_us.amrid;"
SQL="ALTER TABLE Ecoli_IB_AST CHANGE COLUMN `VALUE` `amr_VALUE` longtext;"
SQL="CREATE TABLE IF NOT EXISTS Ecoli_IB_merged AS  SELECT * FROM Ecoli_IB_merged   INNER JOIN Ecoli_IB_AST ON  Ecoli_IB_merged.amrid = Ecoli_IB_AST.ast_id;"
SQL="ALTER TABLE `isolation_source_ecoli_IB` ADD COLUMN `iso_id` INT AUTO_INCREMENT UNIQUE FIRST;"
SQL="CREATE TABLE IF NOT EXISTS EcoliIBisomerge AS  SELECT * FROM Ecoli_IB_merged   INNER JOIN isolation_source_ecoli_IB ON  Ecoli_IB_merged.isoid = isolation_source_ecoli_IB.iso_id;"
SQL=" ALTER TABLE `Date_ecoli_IB` ADD COLUMN `dat_id` INT AUTO_INCREMENT UNIQUE FIRST;"
SQL="CREATE TABLE IF NOT EXISTS EcoliIBiso_dat_merge AS  SELECT * FROM EcoliIBisomerge   INNER JOIN Date_ecoli_IB ON  EcoliIBisomerge.isoid = Date_ecoli_IB.dat_id;"
SQL="ALTER TABLE `Loc_ecoli_IB` ADD COLUMN `loc_id` INT AUTO_INCREMENT UNIQUE FIRST;"
SQL="CREATE TABLE IF NOT EXISTS EcoliIBisodatlocmerge AS  SELECT * FROM EcoliIBiso_dat_Merge   INNER JOIN Loc_ecoli_IB ON  EcoliIBiso_dat_Merge.isoid = Loc_ecoli_IB.loc_id;"
SQL="update EcoliIBisodatlocmerge  set cd_year=replace(cd_year, '2030','Missing');"
SQL="update EcoliIBisodatlocmerge  set state_loc=replace(state_loc, 'Ronald Reagan UCLA Medical Center','California');"
SQL="update EcoliIBisodatlocmerge  set samplehost=replace(samplehost, 'Caninae','Dog');"
SQL="select state_loc,LocationIB  from EcoliIBisodatlocmerge where LocationIB like '%Midwest%';"
SQL="Update EcoliIBisodatlocmerge  set state_loc = 'Unknown' where state_loc is NULL"






#%%


#creating a table where all are scientific name Escherichia Coli 
#%%
new_table= cursor.execute("CREATE TABLE IF NOT EXISTS Ecoli_IB_Merged AS SELECT * FROM Ecoli_IB_merged_AMR  WHERE Scientific_name LIKE '%Escherichia coli%'  ")
print("Selected %s rows" %new_table)
print("Selected %s rows " %cursor.rowcount) 
#rows =cursor.fetchall()#fetch all rows at once
#print(tabulate(rows, headers=['strain','AST_phenotype', 'AMR_genotype'], tablefmt='psql'))

#%%





#writing data table into an excel file
#%%
import pandas as pd
numrows= cursor.execute("Describe Ecoli_IB_Merged")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
print(rows)
dfz= pd.DataFrame(rows)

listz=list(dfz[0])
numrows_k= cursor.execute("Select * from Ecoli_IB_Merged")
print("Selected %s rows" %numrows_k)
print("Selected %s rows " %cursor.rowcount)
rows_k =cursor.fetchall()#fetch all rows at once
df_kleb= pd.DataFrame(rows_k)
df_kleb.columns= listz
#df_kleb.to_excel("Ecoli_IB_merged.xlsx")
#%%


#%%
df_kleb["sourcehostcomb"] = df_kleb["Isolation_Source"].fillna('IMV') + df_kleb["Host"].fillna('HMV')+df_kleb["Isolation_type"].fillna('TMV')


#%%

#
#%%
xdf=pd.crosstab(  df_kleb.sourcehostcomb,df_kleb.Isolation_type,  margins=True, margins_name="total")
#xdf.to_excel("Isolation_source.xlsx")


#%%

#%%
import re
xdf=xdf.reset_index()
xdf["epitype"]=xdf["sourcehostcomb"]
xdf=xdf.replace('')
xdf['epitype'] = xdf['epitype'].map(lambda x: re.sub(r'\W+', '', x)) #removes all characters which are not words 
xdf['epitype'] = xdf['epitype'].map(lambda x: re.sub(" ", "", x))
xdf['epitype'] = xdf['epitype'].str.lower()#converts the colmn into all lower cases
xdf["epi_hostname"]=xdf["epitype"]



#%%

#the sequence to execute the next 5 snippets is essential


#%%

xdf.loc[xdf.epi_hostname.str.contains('cellculturehmvenv|bedding|aquarium|sheeppen|groundhmv|chickenhouse|house|sediment|soil|feedlot|lab|swabmissing|watersource',na=False),'epi_hostname'] = 'Environment:Environment'
#%%

#%%
xdf.loc[xdf.epi_hostname.str.contains('stockpellet|livestockfeed|animalfeed|animalfeed|milksus|milkporcine|petfood|peaproteinpellet|animalfeed|grass|feedgrass|feedhmv|catfood|dogfood',na=False),'epi_hostname'] = 'Food:Animalfood'
xdf.loc[xdf.epi_hostname.str.contains('cheese|chocolate|icecream|dairyproducts|milk|roquefort',na=False),'epi_hostname'] = 'Food:Dairy'
xdf.loc[xdf.epi_hostname.str.contains('veggie|piecrust|lettuce|spinach|coriander|sprouts|pepper|flour|alfalfa|leafygreens|basil|cherrytomato|cucumber|parsley|pizzadough|kale|grain|celery|almond|berries|berry|applecider|apple|cider|chard|cilantro|clover|soynut|tomato|carrot|forage|mint|springmix|vegetable|romaine|melon|microgreen|nuts|organic|oregano|basil|freshsage|rosemary|bean|foodhmv|foodsamplehmv|foodsamplefood|cepa|kolarabi|peaprotein|cabbage|leaf|thyme|springmix|creamysoynutbutter|soynut|cantaloupehmv|cantalopehmv',na=False),'epi_hostname'] = 'Food:plant_based'
xdf.loc[xdf.epi_hostname.str.contains('porkhmv|wholechicken|chickenskin|skinsamplegallus|foodturkey|egg|vealtrim|beefstreak|carcass|carcassswab|beefcarctrim|beefcutlet|beefpatties|patties|turkeypatties|rawveal|rawbeef|beeftrim|groundbeef|groundpork|groundmeat|groundturkey|meatsticks|meatsauce|tenderloin|hamburger|burger|meat|meatloaf|commercial|salami|shrimp|piecrust|venison|boneless|breast|steak|siluriformes|porkchop|product|ground|comminuted|food(pork)|foodpork|packaged|sliced|yolk|chickenwing|chickenthigh|wing|chickenleg|mixedparts|beefcarcass|venison',na=False),'epi_hostname'] = 'Food:Meat/Meatproduct'
#%%



#%%
xdf.loc[xdf.epi_hostname.str.contains('canine|dog|familiaris|familaris',na=False),'epi_hostname'] = 'Companionanimal:Dog'
xdf.loc[xdf.epi_hostname.str.contains('fecesrat|serinuscanaria|pigeon|pelecan|cervuscanadensis|catherpesmexicanus|birdavian|fecesavian|liveravian|bonemarrowavian|imvavian|imvrat|avianhmv|avianmacaw|avianliver|rabbit|deer|raccoon|snake|reptile|boar|commoneider|elk|mouse|grouse|cygnus|quiscalus|beaver|viginianus|caribou|macroura|platyrhynchos|owlpellet|oyster|serpent|rabbit|cuniculus|mink|mus|ratfeces|rathmv|rodent|leucopus|virginianus|branta|capra|hircus|corvus|gull|larus|laridae|crow|cougaranser|crotalus|mephitis|insect|tortoise|myotis|humeralis|hemionus|quail|otter|phocidae|mouflon|moulfon|viridis|passer|colchicus|raccon|otariinae|otariidae|anser|daisypodidae|cricetinae|squirrel|parakeet|galliformes|cougar|flies|canis|lupus|cayote|coyote|prairiedog|ovisorientalis|oviscanadensis',na=False),'epi_hostname'] = 'Clinical:Wildanimal'
xdf.loc[xdf.epi_hostname.str.contains('mesocricetusauratus|ailuropoda|somateria|spectabilis|rhesus|ceratotherium|serinuscanaris|capybaras|accipitridae|gaur|antelope|gazelle|dolphin|leopard|celebeseape|hedgehog|kangaroo|cetacea|duiker|ape|lervia|neovisonvison|camel|stelleri|fischeri|cockatoo|roe|gazella|oreotragus|klipspringer|niger|spekii|gerenuk|elephant|ibex|bontebok|mephitidae|okapi|orangutan|orca|panthera|primate|psitta|rhinoceros|rucervus|girraffe|gorilla|furo|ostrich|panda|muntjac|sika|sabaeus|goral|nippon|wapiti|penguin|nayaur|urial|molinae|unicolor|trichechus|pacos|llama|eider|sirenia|haliaeetus|ararauna|walrus|kea|whale|lagopus|zebra|lotor|ottarinae|ottaridae|dama|marmoset|macaca|kobus|delphinus|crotalus|chlorocebus|lion|monkey|lemur|jacchus|bison|giraffe|ferret|vicugna|cephalophus|cephaloph|dasypodidae|seal|parrot|guineapig',na=False),'epi_hostname'] = 'Clinical:Exotic/Zoo-animal'
xdf.loc[xdf.epi_hostname.str.contains('horse|equus|columba|equine|equis',na=False),'epi_hostname'] = 'Farmanimal:Horse'
xdf.loc[xdf.epi_hostname.str.contains('fecesfeline|lungcat|catfeline|imvcat|cathmv|feliscatus|urinefeline',na=False),'epi_hostname'] = 'Companionanimal:Cat'
#%%




#%%
xdf.loc[xdf.epi_hostname.str.contains('tissuehmvclinical|isolatehmvclinical|strainclinical|isolateclinical|isolatehomosapien|stoolhmvclinical|feceshomo|feceshmvclinical|fecesclinical|fecalsamplehomo|fecalmaterial|fluidhmv|brothstool|bloodculture|cathether|woundsite|tracheal|ulcer|umbilicus|college|surgical|patient|neck|bodyfluid|wounds|groin|sputum|sepsis|respiratory|catheter|drain|aspirate|abscess|genitourinary|stoolsample|jointinfection|human|facility|homosapien|bloodhmv|infant|rectalswab|swabhomosapien|rectal|urinehmv',na=False),'epi_hostname'] = 'Clinical:Homosapiens'
#%%




#%%

xdf.loc[xdf.epi_hostname.str.contains('bullmanure|ovine|bostaurus|beef|cattle|cow|bos|bovine|calf|veal|buffalo|goat|udder|mastitis|sheep|caprahircus|ovisaries',na=False),'epi_hostname'] = 'Farmanimal:Ruminant'
xdf.loc[xdf.epi_hostname.str.contains('pig|porcine|swine|sus|feralswine|feralpig',na=False),'epi_hostname'] = 'Farmanimal:Swine'
xdf.loc[xdf.epi_hostname.str.contains('turkey|chicken|goose|duck|gallus|meleagris',na=False),'epi_hostname'] = 'Farmanimal:Poultry'

#%%




#%%
xdf.loc[xdf.epi_hostname.str.contains('clinical',na=False),'epi_hostname'] = 'Clinical:unknown'


#%%
#%%

xdf.loc[xdf.epi_hostname.str.contains('yard|dairyprocessing|barn|sheeppen|pasture|fecal|feces|colostrum|unknownspecimen|dispenser|door|bench|computer|creek|boot|bedding|dragswab|soappump|chickenhouse|chickenfecalhouse|chickenlitterhouse|chickenwatersource|poultrylitter|dispenser|soil|sand|water|sediment|beansoil|beanwater|cecal|cecum|cecae|cloacae|chair|spectabilis|drawer|powder|enrichment|farm|feceshmv|grimmia|frame|groundhmvenv|poultryhouse|river|feedlot|manure|compliance|earhmv|field|finishedproduct|pump|trap|lairage|litter|livestock|backyard|poultry|wood|produce|rawproduct|tritip|reclaimed|rug|thermometer|run|scope|shoes|upshore|bathroom|swabhmvenv|sink|feedlot|swabhmvenv',na=False),'epi_hostname'] = 'Environment:Environment'

#%%


#%%

xdf.loc[xdf.epi_hostname.str.contains('unknownspecimen|environment|env',na=False),'epi_hostname'] = 'Environment:unknown'

#%%






#%%
df_match=pd.DataFrame()
df_match["isolate_source"]=df_kleb["sourcehostcomb"]
xdf_dict=xdf.set_index('sourcehostcomb')['epi_hostname'].to_dict()
df_match['hostepi'] = df_match['isolate_source'].map(xdf_dict).apply(lambda x: x[0] if type(x)==list else x)
xdf.epi_hostname.value_counts()
df_match.hostepi.value_counts()


#%%

#%%
xdf["epi_type"]=xdf["epi_hostname"]
xdf.loc[xdf.epi_type.str.contains('Clinical|Companion|Farm',na=False),'epi_type'] = 'Clinical'
xdf.loc[xdf.epi_type.str.contains('Environment',na=False),'epi_type'] = 'Environment'
xdf.loc[xdf.epi_type.str.contains('Food',na=False),'epi_type'] = 'Food'
xdf.epi_type.value_counts()


#%%

#%%
xdf_dict=xdf.set_index('sourcehostcomb')['epi_type'].to_dict()
df_match['epi_type'] = df_match['isolate_source'].map(xdf_dict).apply(lambda x: x[0] if type(x)==list else x)
df_match.epi_type.value_counts()
#%%


#%%
xdf["samplehost"]=xdf["epi_hostname"]
xdf["samplehost"]=xdf.samplehost.str.split(':').str[1]

xdf.samplehost.value_counts()


#%%
#%%
xdf_dict=xdf.set_index('sourcehostcomb')['samplehost'].to_dict()
df_match['samplehost'] = df_match['isolate_source'].map(xdf_dict).apply(lambda x: x[0] if type(x)==list else x)
df_match.epi_type.value_counts()
#%%





#%%
xdf["epi_host"]=xdf["epi_hostname"]
xdf.loc[xdf.epi_host.str.contains('animal',na=False),'epi_host'] = 'Animal'
xdf.loc[xdf.epi_host.str.contains('Environment',na=False),'epi_host'] = 'Environment'
xdf.loc[xdf.epi_host.str.contains('Food',na=False),'epi_host'] = 'Food'
xdf.loc[xdf.epi_host.str.contains('Homosapiens|Clinical:unknown',na=False),'epi_host'] = 'Homosapiens'


xdf.epi_host.value_counts()


#%%

#%%
xdf_dict=xdf.set_index('sourcehostcomb')['epi_host'].to_dict()
df_match['epi_host'] = df_match['isolate_source'].map(xdf_dict).apply(lambda x: x[0] if type(x)==list else x)
df_match.epi_host.value_counts()
#%%

#%%

xdf.to_excel("Isolation_source_dictionary1.xlsx")


#%%



#run this code snippet just once to create the final table
#%%
from sqlalchemy import create_engine
cnx = create_engine('mysql+pymysql://root:Sukhoi@90@localhost/myamr')
df_match.to_sql(name='isolation_source_ecoli_IB', con=cnx, if_exists = 'replace', index=False)
print("we are at line 244 in code")

#
#
#%%





#writing data table into an excel file
#%%
import pandas as pd
numrows= cursor.execute("Describe EcoliIBisodatlocMerge")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
print(rows)
dfz= pd.DataFrame(rows)

listz=list(dfz[0])
numrows_k= cursor.execute("Select * from EcoliIBisodatlocMerge")
print("Selected %s rows" %numrows_k)
print("Selected %s rows " %cursor.rowcount)
rows_k =cursor.fetchall()#fetch all rows at once
df_kleb= pd.DataFrame(rows_k)
df_kleb.columns= listz
#df_kleb.to_excel("EcoliIBIsomerged.xlsx")
#%%

#%%

df_kleb.to_excel("Ecoliibmergedisodatloc.xlsx")


#%%


#we need to convert the dataframe into a series of list to use the strftime() function as shown in the reference example
#%%
numrows_k= cursor.execute("Select Collection_Date from EcoliIBisoMerge ;")
print("Selected %s rows" %numrows_k)
print("Selected %s rows " %cursor.rowcount)
rows_k =cursor.fetchall()#fetch all rows at once
df_CD= pd.DataFrame(rows_k)      
df_CD = df_CD.fillna('2030') 
df_CD = df_CD.replace('1900/1949','1949') 
df_CD = df_CD.replace('1949/1900','1949')
df_CD = df_CD.replace('1900/1960','1960')
df_CD = df_CD.replace('2017/2018','2018')
df_CD = df_CD.replace('2015-01/2016-07','2015')
df_ser=pd.Series(df_CD.values.tolist(), index=df_CD.index)
df_CD['cd_mod']=df_ser.apply(lambda x: pd.to_datetime(x).strftime('%m-%d-%Y')[0])
df_CD['cd_year']=df_ser.apply(lambda x: pd.to_datetime(x).strftime('%Y')[0])
df_CD=df_CD.rename(columns = {0:'cd_table'})
#%%
#the reference example to convert datetime objects
#%%
df = pd.DataFrame([
    [['']],
    [['May 23rd, 2011']],
    [['January 1st, 2010']],
    [['Apr. 15, 2008']],
    [['07-11-2013']],
    [['9/01/1995']],
    [['04/15/2000']],
    [['11/22/68']],
    [['12/1997']],
    [['08/2002']],
    [['2014']],
    [['2016']],
], columns=['date'])

df['clean_date'] = df.date.apply(
    lambda x: pd.to_datetime(x).strftime('%m-%d-%Y')[0])        
    
print(df)  
    
    
    
#%%

#run this code snippet just once to create the final table
#%%
from sqlalchemy import create_engine
cnx = create_engine('mysql+pymysql://root:Sukhoi@90@localhost/myamr')
df_CD.to_sql(name='Date_ecoli_IB', con=cnx, if_exists = 'replace', index=False)
print("we are at line 244 in code")

#
#
#%%

#%%




#writing data table into an excel file
#%%
import pandas as pd
numrows= cursor.execute("Describe EcoliIBiso_dat_merge")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
print(rows)
dfz= pd.DataFrame(rows)

listz=list(dfz[0])
numrows_k= cursor.execute("Select * from EcoliIBiso_dat_merge")
print("Selected %s rows" %numrows_k)
print("Selected %s rows " %cursor.rowcount)
rows_k =cursor.fetchall()#fetch all rows at once
df_kleb= pd.DataFrame(rows_k)
df_kleb.columns= listz
#df_kleb.to_excel("EcoliIBIsomerged.xlsx")
#%%

#dictionary for city state reference
#%%
import pandas as pd
numrows= cursor.execute("SELECT  City, State FROM latlon_states")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
#print(tabulate(rows, headers=['AST_phenotype'], tablefmt='psql'))
df_loc = pd.DataFrame(rows)
df_loc.columns=["state","statecode"]
loc_dict=df_loc.set_index('statecode')['state'].to_dict()
stcode_list=df_loc['statecode'].tolist()
st_list=df_loc['state'].tolist()
#%%

#get the data from the Ecoli_merged table to make the dataframe for unique locations
#%%
import pandas as pd
numrows= cursor.execute("select  Location from EcoliIBiso_dat_merge group by Location;")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
#print(tabulate(rows, headers=['AST_phenotype'], tablefmt='psql'))
gloc_df = pd.DataFrame(rows)
gloc_df.columns=["Location"]

#%%




#%%
gloc_df["state"]=gloc_df["Location"]
gloc_df["state"]=gloc_df.state.str.split(':').str[1]
gloc_df = gloc_df.fillna('USA')
gloc_df["state"] = gloc_df.state.str.lstrip() #
gloc_df.state.value_counts()
#
#%%

#%%
pat = r'({})'.format('|'.join(stcode_list))
gloc_df['state'] = gloc_df['state'].str.extract(pat, expand=False).fillna(gloc_df['state'])
gloc_df['state'] =gloc_df['state'].map(loc_dict).fillna(gloc_df['state'])
pat = r'({})'.format('|'.join(st_list))
gloc_df['state'] = gloc_df['state'].str.extract(pat, expand=False).fillna(gloc_df['state'])
#%%

#%%
        
        
#%%


gloc_df=gloc_df.replace(to_replace ="Fort Sam Houston", value ="Texas")
gloc_df=gloc_df.replace(to_replace ="Houston", value ="Texas")
gloc_df=gloc_df.replace(to_replace ="Baltimore", value ="Maryland")
gloc_df=gloc_df.replace(to_replace ="St. Louis Clyde watershed of Lake Superior", value ="St Louis")
gloc_df=gloc_df.replace(to_replace ="UC Davis Medical Center", value ="California")
gloc_df=gloc_df.replace(to_replace ="UC Davis Medical Center, Davis, Ca", value ="California")
gloc_df=gloc_df.replace(to_replace ="University of Miami Department of Pathology", value ="Florida")
gloc_df=gloc_df.replace(to_replace ="Burlington", value ="Vermont")
gloc_df=gloc_df.replace(to_replace ="Lahey Hospital & Medical Center", value ="Massachusetts")
gloc_df=gloc_df.replace(to_replace ="Ronald Reagan UCLA Medical Center", value ="California")
gloc_df=gloc_df.replace(to_replace ="Ronald Reagan UCLA Medical Center", value ="Unknown")
#gloc_df.to_excel(gloc.xlsx)

#%% 




#%%
#gloc_df.to_excel("Gloc.xlsx")
loc_dict_df = pd.read_excel('/Users/jha/Gloc.xlsx', sheet_name='Sheet1')
list_loc=list(loc_dict_df.itertuples(index=False, name=None))
#list_family=list(df_family)
loc_dict=dict(list_loc)

#%%

#get the data from the Ecoli_merged table to make the dataframe for unique locations
#%%
import pandas as pd
numrows= cursor.execute("select  Location from EcoliIBiso_dat_merge;")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
#print(tabulate(rows, headers=['AST_phenotype'], tablefmt='psql'))
loc_IB_df = pd.DataFrame(rows)
loc_IB_df.columns=["LocationIB"]

#%%
#%%
loc_IB_df['state_loc']=loc_IB_df['LocationIB'].map(loc_dict) #the location needs to be mapped to state 

#%%

#run this code snippet just once to create the final table
#%%
from sqlalchemy import create_engine
cnx = create_engine('mysql+pymysql://root:Sukhoi@90@localhost/myamr')
loc_IB_df.to_sql(name='Loc_ecoli_IB', con=cnx, if_exists = 'replace', index=False)
print("we are at line 244 in code")

#
#
#%%


#writing data table into an excel file
#%%
import pandas as pd
numrows= cursor.execute("Describe EcoliIBisodatlocmerge")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
print(rows)
dfz= pd.DataFrame(rows)

listz=list(dfz[0])
numrows_k= cursor.execute("Select * from EcoliIBisodatlocmerge")
print("Selected %s rows" %numrows_k)
print("Selected %s rows " %cursor.rowcount)
rows_k =cursor.fetchall()#fetch all rows at once
df_kleb= pd.DataFrame(rows_k)
df_kleb.columns= listz
#%%
#%%


df_kleb.to_excel("EcoliIBIsodatlocmerged.xlsx")
#%%







#%%
import pandas as pd
numrows= cursor.execute("Describe Ecoli_IB_amr")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
print(rows)
df_col_name_amr= pd.DataFrame(rows)
list_col_name_amr=list(df_col_name_amr[0])

del list_col_name_amr[0]
del list_col_name_amr[0]


#%%


#%%
import pandas as pd
numrows= cursor.execute("Describe Ecoli_IB_AST ")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
print(rows)
df_col_name_ast= pd.DataFrame(rows)
list_col_name_ast=list(df_col_name_ast[0])

del list_col_name_ast[0]
del list_col_name_ast[0]


#%%

#for only those antimicrobials which were tested for more than one epitypes
#%%
 #def crosstab_all(df_kleb,list_col_name):
try:
    appended_data=[]
    for k in list_col_name_amr:
        try:
            xdf_crosstab=pd.crosstab(df_kleb["epi_host"],df_kleb[k], margins=True, margins_name="Out of total samples")
            xdf_crosstab["AMR_Gene"] = k
            print('xdf',xdf_crosstab)
            print('') # for spacing
            if len(xdf_crosstab.index) >2:
                appended_data.append(xdf_crosstab)
            else:
                print("only one epitype tested")
                 
        except:
            print("not found")
            continue
    appended_data= pd.concat(appended_data, sort= False) 
    appended_data=appended_data.drop("missing")
    appended_data=appended_data.drop("")
    #appended_data_P=appended_data_P.drop("DD", axis=1)
except:
    print("some error in gene",k)
    

#%%



#%%



appended_data = appended_data.reset_index()
appended_data= (appended_data.groupby([ 'AMR_Gene','epi_host']).sum())

cols=["HMM","PARTIAL","POINT","PARTIAL_END_OF_CONTIG","COMPLETE","MISTRANSLATION","DNE","Out of total samples"]
appended_data =appended_data[cols]
 

#%%

#%%

appended_data.to_excel("crosstabepihost_gene.xlsx")


#%%



#for only those antimicrobials which were tested for more than one epitypes
#%%
 #def crosstab_all(df_kleb,list_col_name):
try:
    appended_data=[]
    for k in list_col_name_amr:
        try:
            xdf_crosstab=pd.crosstab([df_kleb["cd_year"],df_kleb["epi_host"]],df_kleb[k], margins=True, margins_name="Out of total samples")
            xdf_crosstab["AMR_gene"] = k
            print('xdf',xdf_crosstab)
            print('') # for spacing
            if len(xdf_crosstab.index) >0:
                appended_data.append(xdf_crosstab)
            else:
                print("only one epitype tested")
                 
        except:
            print("not found")
            continue
    appended_data= pd.concat(appended_data, sort= False) 
    appended_data=appended_data.drop("missing")
    appended_data=appended_data.drop("")
    #appended_data_P=appended_data_P.drop("DD", axis=1)
except:
    print("some error in gene",k)
    

#%%


#%%



appended_data = appended_data.reset_index()
appended_data= (appended_data.groupby([ 'AMR_gene','cd_year','epi_host']).sum())

cols=["HMM","PARTIAL","POINT","PARTIAL_END_OF_CONTIG","COMPLETE","MISTRANSLATION","DNE","Out of total samples"]
appended_data =appended_data[cols]
 

#%%

#%%
appended_data.to_excel("crosstab_gene_dat_host.xlsx")

#%%


#for only those antimicrobials which were tested for more than one epitypes
#%%
 #def crosstab_all(df_kleb,list_col_name):
try:
    appended_data=[]
    for k in list_col_name_amr:
        try:
            xdf_crosstab=pd.crosstab(df_kleb["samplehost"],df_kleb[k], margins=True, margins_name="z_Out of total samples")
            xdf_crosstab["AMR_gene"] = k
            print('xdf',xdf_crosstab)
            print('') # for spacing
            if len(xdf_crosstab.index) >0:
                appended_data.append(xdf_crosstab)
            else:
                print("only one epitype tested")
                 
        except:
            print("not found")
            continue
    appended_data= pd.concat(appended_data, sort= False) 
    appended_data=appended_data.drop("missing")
    appended_data=appended_data.drop("")
    #appended_data_P=appended_data_P.drop("DD", axis=1)
except:
    print("some error in gene",k)
    

#%%


#%%



appended_data = appended_data.reset_index()
appended_data= (appended_data.groupby([ 'AMR_gene','samplehost']).sum())

cols=["HMM","PARTIAL","POINT","PARTIAL_END_OF_CONTIG","COMPLETE","MISTRANSLATION","DNE","z_Out of total samples"]
appended_data =appended_data[cols]
 

#%%

#%%
appended_data.to_excel("crosstab_AMRgene_samplehost.xlsx")
 
#%%



#for ACRF
#%%
 #def crosstab_all(df_kleb,list_col_name):

xdf_crosstab=pd.crosstab([df_kleb["cd_year"],df_kleb["state_loc"], df_kleb["epi_host"]],df_kleb["acrF"], margins=True, margins_name="Out of total samples")
print('xdf',xdf_crosstab)
    
appended_data=xdf_crosstab

#%%


#%%



appended_data = appended_data.reset_index()
appended_data= (appended_data.groupby(['cd_year','state_loc','epi_host']).sum())

cols=["PARTIAL","PARTIAL_END_OF_CONTIG","COMPLETE","MISTRANSLATION","Out of total samples"]
appended_data =appended_data[cols]
 

#%%



#%%
appended_data.to_excel("crosstab_acrF_dat_loc_epi.xlsx")

#%%


#for ACRF
#%%
 #def crosstab_all(df_kleb,list_col_name):

xdf_crosstab=pd.crosstab([df_kleb["cd_year"],df_kleb["state_loc"], df_kleb["epi_host"]],df_kleb["emrD"], margins=True, margins_name="Out of total samples")
print('xdf',xdf_crosstab)
    
appended_data=xdf_crosstab

#%%


#%%



appended_data = appended_data.reset_index()
appended_data= (appended_data.groupby(['cd_year','state_loc','epi_host']).sum())

cols=["PARTIAL","PARTIAL_END_OF_CONTIG","COMPLETE","Out of total samples"]
appended_data =appended_data[cols]
 

#%%



#%%
appended_data.to_excel("crosstab_emrD_dat_loc_epi.xlsx")

#%%


#for only those antimicrobials which were tested for more than one epitypes
#%%
 #def crosstab_all(df_kleb,list_col_name):
try:
    appended_data=[]
    for k in list_col_name_amr:
        try:
            for ast in list_col_name_ast:
                xdf_crosstab=pd.crosstab(df_kleb[ast],df_kleb[k], margins=True, margins_name="z_Out of total samples")
                xdf_crosstab["AMR_gene"] = k
                xdf_crosstab["Antimicrobial"]=ast
                print('xdf',xdf_crosstab)
                print('') # for spacing
                if len(xdf_crosstab.index) >0:
                    appended_data.append(xdf_crosstab)
                else:
                    print("only one epitype tested")
                 
        except:
            print("not found")
            continue
    appended_data= pd.concat(appended_data, sort= False) 
    appended_data=appended_data.drop("missing")
    appended_data=appended_data.drop("")
    #appended_data_P=appended_data_P.drop("DD", axis=1)
except:
    print("some error in gene",k)
    

#%%


#%%



appended_data = appended_data.reset_index()
appended_data= appended_data.rename(columns={"index":"Phenotype"})
appended_data= (appended_data.groupby([ 'AMR_gene','Antimicrobial','Phenotype']).sum())

cols=["PARTIAL","POINT","PARTIAL_END_OF_CONTIG","COMPLETE","MISTRANSLATION","z_Out of total samples"]
appended_data =appended_data[cols]
 

#%%

#%%
appended_data.to_excel("crosstab_amr_ast.xlsx")
#%%






#get the data from the Ecoli_merged table to make the dataframe for unique locations
#%%
import pandas as pd
numrows= cursor.execute("select  Location,count(Location) from Ecoli_emrD group by Location;")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
#print(tabulate(rows, headers=['AST_phenotype'], tablefmt='psql'))
loc_df = pd.DataFrame(rows)
loc_df.columns=["Location","Frequency"]
loc_df.to_excel("ecoli_location.xlsx")
#%%


#get the data from the Ecoli_merged table to make the dataframe for unique locations
#%%
import pandas as pd
numrows= cursor.execute("select collection_date, count(collection_date) from EcoliIBisodatlocmerge group by collection_date ")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
#print(tabulate(rows, headers=['AST_phenotype'], tablefmt='psql'))
dat_df = pd.DataFrame(rows)
dat_df.columns=["Location","Frequency"]
dat_df.to_excel("ecoli_date.xlsx")
#%%

#get the data from the Ecoli_merged table to make the dataframe for unique locations
#%%
import pandas as pd
numrows= cursor.execute("select isolate_source, samplehost from EcoliIBisodatlocmerge group by isolate_source;")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
#print(tabulate(rows, headers=['AST_phenotype'], tablefmt='psql'))
loc_df = pd.DataFrame(rows)
loc_df.columns=["Isolation_Source","Samplehost"]
loc_df.to_excel("isolation_samplehost.xlsx")
#%%



#get the data from the Ecoli_merged table to make the dataframe for unique locations
#%%
import pandas as pd
numrows= cursor.execute("select  Isolation_type,count(Isolation_type) from Ecoli_IB_Merged group by Isolation_type;")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
#print(tabulate(rows, headers=['AST_phenotype'], tablefmt='psql'))
loc_df = pd.DataFrame(rows)
loc_df.columns=["Isolation_type","Frequency"]
loc_df.to_excel("ecoli.xlsx")
#%%

#get the data from the Ecoli_merged table to make the dataframe for unique locations
#%%
import pandas as pd
numrows= cursor.execute("select  Host,count(Host) from Ecoli_IB_merged_AMR group by Host;")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
#print(tabulate(rows, headers=['AST_phenotype'], tablefmt='psql'))
loc_df = pd.DataFrame(rows)
loc_df.columns=["Host","Frequency"]
loc_df.to_excel("ecoli_location.xlsx")
#%%

#get the data from the Ecoli_merged table to make the dataframe for unique locations
#%%
import pandas as pd
numrows= cursor.execute("ALTER TABLE EcoliIBisodatlocmerged DROP COLUMN samplehost;")
#print("Selected %s rows" %numrows)
print("DONE")

#%%




































##%%
#from sqlalchemy import create_engine
#
#cnx = create_engine('mysql+pymysql://root:Sukhoi@90@localhost/myamr')
##df3.to_sql(name='salall', con=cnx, if_exists = 'replace', index=False)
#df_sal_ast= pd.read_sql('Klebsiellaast',cnx)
#
##%%
#
#
#
#
#%%

numrows= cursor.execute("CREATE TABLE IF NOT EXISTS Ecoli_nov_merged AS SELECT * FROM Ecoli_nov_ast INNER JOIN Ecoli_nov_astt ON Ecoli_nov_ast.AST_phenotypes = Ecoli_nov_astt.AST")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
#print(tabulate(rows, headers=['strain','AST_phenotype', 'AMR_genotype'], tablefmt='psql'))    
#dffinal = pd.DataFrame(rows) 

#%%
#
#
#
#%%
#header_astt=[]
#header_astt= list(df_ASTT.columns)
#header_kleb_ast=[]
#header_kleb_ast= list(df_kleb.columns)
#header_final=[]
#header_final=header_astt+header_kleb_ast
#print(header_final)
#header_final = [s.replace('#label', 'Label') for s in header_final]
#dffinal.columns=header_final
#dffinal=dffinal.drop(dffinal.columns[[0]], axis=1)
##
#
#
#
#select * from Persons inner join data on persons.ID = dat
#
#
#%%
#
#
#
#run this code snippet just once to create the final table
#%%
from sqlalchemy import create_engine
cnx = create_engine('mysql+pymysql://root:Sukhoi@90@localhost/myamr')
dffinal.to_sql(name='Klebsiella_all_us', con=cnx, if_exists = 'replace', index=False)
print("we are at line 244 in code")

#
#
#%%
#

#this is the same as above why am I doing this again??
#%%
df_AB=df_AST
df_AB= df_AB.T
df_AB.columns= df_AB.iloc[0]
df_AB=df_AB.drop([0])
df_AB=df_AB.reset_index()
df_AB=df_AB.drop(columns=["index"])

#%%

#This removes all drugs which have only none values in all rows and columns 
#%%

df_AB = df_AB.replace('ND', np.nan)
df_AB = df_AB.fillna('0')
df_AB = df_AB.replace('0', np.nan)
#df_AB = df_AB.set_index('AST_phenotypes')
df_AB = df_AB.dropna(axis = 0, how = 'all')
df_AB = df_AB.dropna(axis = 1, how = 'all')
#df_AB = df_AB.reset_index()

#%%


#%%
chunk=[]
chunk1=[]
for row in df_AB.columns:
    try:
        print(row)
        #print(df_AB[row].value_counts().index.to_list())
        #df_plot=(df_AB[row].value_counts().index.to_list())
        
        #df_plot=pd.merge((df_AB[row].value_counts().to_frame()), on )
        #print(df_AB[row].value_counts().to_frame())
        #df_plot= df.append(df_AB[row].value_counts(), ignore_index=True)
        #print("line 312")
        
        print(df_AB[row].value_counts().reset_index())
        df5=df_AB[row].value_counts().reset_index()
        val=df5.loc[df5['index'] != "S"]
        val=val.T
        val=val.reset_index()
        val=val.drop([0])
        print("before saving to list)")
        chunk1.append(val.values.tolist())
        #val.columns=['antibiotic','number_of_S']
        
        
        
        
        
        #df=pd.DataFrame('Antimicrobial': val[])
    #    print("yes",val)
        #chunk= [df5.columns.values.tolist()] + df5.values.tolist()
    #    df_plot_C=pd.concat(series5)
    except:
        pass
    

    
#%%


    
#%%
list_chunk=[]       
for x in chunk1:
            print(x)
            for y in x:
                print(y)
                list_chunk.append(y)    
    
    
#%%        


                
        
#%%
        
from matplotlib import pyplot as plt
import matplotlib.cm as cm
from matplotlib.colors import Normalize
import pandas as pd
import numpy as np
df_plotKNS= pd.DataFrame(list_chunk)

df_plotKNS.columns=["Antimicrobial","R","I"] #when there is all I in a column or all R then the I column is included in the R column
#df_plotKNS=df_plotKNS.drop(columns=["ND"])

df_plotKNS["non_susceptible"]=df_plotKNS.fillna(0)["R"]+df_plotKNS.fillna(0)["I"] # this line is needed to make non susceptible column by adding R and I columns 
#df_plot.columns=['Antimicrobial','number_of_S']
#%%
                
#percentage resistant to each drug in klebsiella data
#%%
df_plotKNS['percent_nonS'] = ((df_plotKNS['non_susceptible']/(000000))*100) 
df_plotKNS.percent_nonS=df_plotKNS.percent_nonS.round(2)

#df_plotKNS_sorted=df_plotKNS.sort_values('percent_nonS',ascending=False)
#ax = df_plotKNS_sorted.plot.bar(x='Antimicrobial', y='percent_nonS', rot=90,legend=False,    # Turn the Legend off
#        width=0.75,      # Set bar width as 75% of space available
#        figsize=(15,6),  # Set size of plot in inches
#        colormap='summer')
#for index,data in enumerate(df_plotKNS_sorted['percent_nonS']):
#    plt.text(x=index , y =data , s=f"{data}",fontdict=dict(fontsize=8) )
#plt.title("Percentage of Klebsiella samples found non Susceptible to Antimicrobial (33,222 samples)")
#plt.ylabel("percentage of nonS")
#plt.plot(x,y)
#plt.show()




#%%                


#for graphing the susceptible microbes
#%%
        
chunk2=[]
chunk3=[]
for row in df_AB.columns:
    try:
        print(row)
        #print(df_AB[row].value_counts().index.to_list())
        #df_plot=(df_AB[row].value_counts().index.to_list())
        
        #df_plot=pd.merge((df_AB[row].value_counts().to_frame()), on )
        #print(df_AB[row].value_counts().to_frame())
        #df_plot= df.append(df_AB[row].value_counts(), ignore_index=True)
        #print("line 312")
        
        print(df_AB[row].value_counts().reset_index())
        df6=df_AB[row].value_counts().reset_index()
        val=df6.loc[df6['index'] == "S"]
        val=val.T
        val=val.reset_index()
        val=val.drop([0])
        print("before saving to list)")
        chunk3.append(val.values.tolist())
        #val.columns=['antibiotic','number_of_S']
        
        
        
        
        
        #df=pd.DataFrame('Antimicrobial': val[])
    #    print("yes",val)
        #chunk= [df5.columns.values.tolist()] + df5.values.tolist()
    #    df_plot_C=pd.concat(series5)
    except:
        pass   
    
    
    
    
    
    
#%%
    
    

    
#%%
list_chunkS=[]       
for x in chunk3:
            print(x)
            for y in x:
                print(y)
                list_chunkS.append(y)    
    
    
#%%
        
#%%
        
from matplotlib import pyplot as plt
import matplotlib.cm as cm
from matplotlib.colors import Normalize
import pandas as pd
import numpy as np
df_plotKS= pd.DataFrame(list_chunkS)

df_plotKS.columns=["Antimicrobial","Susceptible"]

##                           
##print()
#%%




##plot for % susceptible Antimicrobial to klebsiella Antimicrobial
#%%
#df_plotKS['percent_S'] = ((df_plotKS['Susceptible']/16680)*100) 
#df_plotKS.percent_S=df_plotKS.percent_S.round(2)
#df_plotKS_sorted=df_plotKS.sort_values('percent_S',ascending=False)
#ax = df_plotKS_sorted.plot.bar(x='Antimicrobial', y='percent_S', rot=90,legend=False,    # Turn the Legend off
#        width=0.75,      # Set bar width as 75% of space available
#        figsize=(15,6),  # Set size of plot in inches
#        colormap='summer')
#for index,data in enumerate(df_plotKS_sorted['percent_S']):
#    plt.text(x=index , y =data , s=f"{data}",fontdict=dict(fontsize=8) )
#plt.title("Percentage of Klebsiella samples found Susceptible to Antimicrobial (33222 samples)")
#plt.ylabel("percentage of S")
#plt.plot(x,y)
#plt.show()
#
#
#
#
#%%


#making the table with each percentage non-susceptible is non-susceptible/(susceptible+non_susceptible) same for susceptible

#%%

df_plotKS_dict=df_plotKS.set_index('Antimicrobial')['Susceptible'].to_dict()

df_plotKNS['Susceptible'] = df_plotKNS['Antimicrobial'].map(df_plotKS_dict)
df_plotKNS["Total_Samples"]=df_plotKNS.fillna(0)["Susceptible"]+ df_plotKNS.fillna(0)["non_susceptible"]

df_plotKNS['percent_S'] = ((df_plotKNS['Susceptible']/df_plotKNS["Total_Samples"])*100) 


df_plotKNS['percent_nonS'] = ((df_plotKNS['non_susceptible']/df_plotKNS["Total_Samples"])*100) 





#%%

#Use only if plotting the ib=ncreasing percentage in graphs
#%%
df_plotKNS_sorted=df_plotKNS.sort_values('percent_S',ascending=False)
df_plotKNS_sorted.percent_nonS=df_plotKNS.percent_S.round(0)
df_plotKNS_sorted.percent_nonS=df_plotKNS.percent_nonS.round(0)


#%%




#percentage plot for increasing percentage
#%%

import matplotlib.pyplot as plt
from matplotlib import rc
import numpy as np
import matplotlib.patches as mpatches
r=[]
r= list(df_plotKNS_sorted["Antimicrobial"])
bars1=[]

bars1= list(df_plotKNS_sorted.fillna(0)["percent_nonS"])

bars2=[]

bars2=list(df_plotKNS_sorted.fillna(0)["percent_S"])
p1=plt.bar(r, bars1, color='rosybrown', edgecolor='white', width=1,label="non-Susceptible")
# Create green bars (middle), on top of the firs ones
p2=plt.bar(r, bars2, bottom=bars1, color='mistyrose', edgecolor='white', width=1,label="Susceptible")
plt.xticks(r,rotation=90)
plt.xlabel("Antimicrobial")
plt.legend(loc='upper left', bbox_to_anchor=(1,1), ncol=1)
plt.title("Percentage of Klebsiella samples tested ( samples)")
plt.ylabel("percentage")

rects1 = p1.patches
labels1 = ["%d" % i for i in (df_plotKNS_sorted.fillna(0)["Susceptible"])]
for rect, label in zip(rects1, labels1):
    height = rect.get_height()
    plt.text(rect.get_x() + rect.get_width() / 2., height, label,ha='center', va='bottom',color="black", fontsize=8,fontweight="bold", rotation= 90)


rects2 = p2.patches
labels2 = ["%d" % i for i in df_plotKNS_sorted.fillna(0)["non_susceptible"]]
for rect, label in zip(rects2, labels2):
    height = rect.get_height()
    plt.text(rect.get_x() + rect.get_width() / 2., 100-height, label,ha='center', va='top',color="blue", fontsize=8,fontweight="bold", rotation= 90)


rects1 = p1.patches
labels3 = ["%d" % i for i in df_plotKNS_sorted.fillna(0)["Total_Samples"]]
for rect, label in zip(rects2, labels3):
    height= 20
    plt.text(rect.get_x() + rect.get_width() / 2.,height, label,ha='center', va='bottom',color="maroon", fontsize=8,fontweight="bold", rotation= 90)


plt.show()





#%%






#writing data table into an excel file
#%%
import pandas as pd
numrows= cursor.execute("Describe Escherichia_MergedF")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
print(rows)
dfz= pd.DataFrame(rows)

listz=list(dfz[0])
numrows_k= cursor.execute("Select * from Escherichia_MergedF")
print("Selected %s rows" %numrows_k)
print("Selected %s rows " %cursor.rowcount)
rows_k =cursor.fetchall()#fetch all rows at once
df_kleb= pd.DataFrame(rows_k)
df_kleb.columns= listz

#%%


#This removes all drugs which have only none values in all rows and columns 
#%%

df_kleb = df_kleb.replace('ND', np.nan)
df_kleb = df_kleb.replace('NONE',np.nan)
df_kleb = df_kleb.fillna('0')
df_kleb = df_kleb.replace('0', np.nan)
#df_kleb = df_kleb.set_index('AST_phenotypes')
df_kleb = df_kleb.dropna(axis = 0, how = 'all')
df_kleb = df_kleb.dropna(axis = 1, how = 'all')
#df_kleb = df_kleb.reset_index()

#%%


#run this code snippet just once to create the final table
#%%
from sqlalchemy import create_engine
cnx = create_engine('mysql+pymysql://root:Sukhoi@90@localhost/myamr')
df_kleb.to_sql(name='Escherichia_isolationepi', con=cnx, if_exists = 'replace', index=False)
print("we are at line 244 in code")

#
#
#%%



#%%
#import os
#
#outname = 'klebsiella.csv'
#
#outdir = '/Users/jha/Documents/Fall2019'
#if not os.path.exists(outdir):
#    os.mkdir(outdir)
#
#fullname = os.path.join(outdir, outname)    
#dffinal.to_excel("Klebsiella_all_us.xlsx")

df_kleb.to_excel("EscherichiamergedUS.xlsx")


#%%




#%%

#writing the list of drugs into an excel file for classification

# import xlsxwriter module 
#import xlsxwriter 
#  
#workbook = xlsxwriter.Workbook('klebsiella_full.xlsx') 
#worksheet = workbook.add_worksheet() 
#  
## Start from the first cell. 
## Rows and columns are zero indexed. 
#row = 0
#column = 0
#  
#content = df_kleb
#  
## iterating through content list 
#for item in content : 
#  
#    # write operation perform 
#    worksheet.write(row, column, item) 
#  
#    # incrementing the value of row by one 
#    # with each iteratons. 
#    row += 1
#      
#workbook.close() 











#%%




#**************************************************************************************************


"""
Exploratory statistics on klebsiella

"""




#****************************************************************************************************



#%%
import csv
import mysql.connector as mariadb
from tabulate import tabulate
import pandas as pd

amrdb= mariadb.connect(
        host="localhost",
        user="root",
        passwd="Sukhoi@90",
        database ="myamr"
        )
cursor = amrdb.cursor(buffered=True) # else it fetches one row for everytime it is executed 
print("We are at line 31 we have connection, lets begin")

#%%


#%%

df_family = pd.read_excel('/Users/jha/Documents/spring2020/Klebsiella_Sneha_Jan_2020.xlsx', sheet_name='Sheet1')
df_family_mod= df_family
df_family=df_family.drop(["Intrinsicly_resistant","Antimicrobial_class","Antimicrobial_mod"], axis=1)
list_family=list(df_family.itertuples(index=False, name=None))
#list_family=list(df_family)
family_dict=dict(list_family)
df_plotKNS['Group']=df_plotKNS['Antimicrobial'].map(family_dict)
df_plotKNS_sorted=df_plotKNS.sort_values(by=['Group'])
df_plotKNS_sorted_ref=df_plotKNS_sorted

#%%

#%%
#df_plotKNS_sorted=df_plotKNS_sorted.set_index(['Antimicrobial_class', 'Antimicrobial'])

#%%


#%%

import matplotlib.pyplot as plt
from matplotlib import rc
import numpy as np
import matplotlib.patches as mpatches
r=[]
r2=[]
r= list(df_plotKNS_sorted.Antimicrobial.str.title())
#r = [sr.capitalize() for sr in r]
r2=list(df_plotKNS_sorted["Group"])
bars1=[]

bars1= list(df_plotKNS_sorted.fillna(0)["percent_nonS"])

bars2=[]

bars2=list(df_plotKNS_sorted.fillna(0)["percent_S"])
p1=plt.bar(r,bars1, color='rosybrown', edgecolor='white', width=1,label="Non-susceptible")
# Create green bars (middle), on top of the firs ones
p2=plt.bar(r, bars2, bottom=bars1, color='mistyrose', edgecolor='white', width=1,label="Susceptible")
plt.xticks(r,rotation=90)
plt.xlabel("Antimicrobials")
plt.legend(loc='upper left', bbox_to_anchor=(1,1), ncol=1)
plt.title("Comparative percentage of Klebsiella samples tested in 2015,grouped by family of antimicrobials")
plt.ylabel("Percentage")






rects1 = p1.patches
labels1 = ["%d" % i for i in (df_plotKNS_sorted.fillna(0)["non_susceptible"])]
for rect, label in zip(rects1, labels1):
    height = rect.get_height()
    plt.text(rect.get_x() + rect.get_width() / 2., height-1, label,ha='center', va='top',color="blue", fontsize=8,fontweight="bold", rotation= 90)


rects2 = p2.patches
labels2 = ["%d" % i for i in df_plotKNS_sorted.fillna(0)["Susceptible"]]
for rect, label in zip(rects2, labels2):
    height = rect.get_height()
    plt.text(rect.get_x() + rect.get_width() / 2., 100-height, label,ha='center', va='bottom',color="black", fontsize=8,fontweight="bold", rotation= 90)


rects1 = p1.patches
labels3 = ["%d" % i for i in df_plotKNS_sorted.fillna(0)["Total_Samples"]]
for rect, label in zip(rects2, labels3):
    height= 0
    plt.text(rect.get_x() + rect.get_width() / 2.,height, label,ha='center', va='bottom',color="black", fontsize=8,fontweight="bold", rotation= 90)



plt.show()





#%%


#%%
print(df_plotKNS_sorted.Antimicrobial_class.value_counts())
listp=[]
listp=list(df_plotKNS_sorted.Antimicrobial_class.unique())
#df_plotKNS_sorted_multi=df_plotKNS_sorted
df_plotKNS_sorted=df_plotKNS_sorted.set_index(['Antimicrobial_class', 'Antimicrobial'])
df_plotKNS_sorted_multi=df_plotKNS_sorted.sum(level =0)


#%%


#%%
# Libraries
import matplotlib.pyplot as plt
 
# Make data: I have 3 groups and 7 subgroups

group_names=listp
group_size=[]
group_size=list(df_plotKNS_sorted_multi["non_susceptible"])
subgroup_names=[]
subgroup_names=list(df_plotKNS_sorted_ref["Antimicrobial"])
subgroup_size=[]
subgroup_size=list(df_plotKNS_sorted_ref["non_susceptible"])
 
# Create colors
#a, b, c=[plt.cm.Blues, plt.cm.Reds, plt.cm.Greens]




cmap = plt.get_cmap('PuBuGn')
colors_grp = cmap(np.linspace(0., 1., len(group_names)))
cmap = plt.get_cmap('PuBuGn')
colors= cmap(np.linspace(0., 1., len(subgroup_names))) 
print(colors_grp)
print(colors)
 #First Ring (outside)
fig, ax = plt.subplots()
ax.axis('equal')
mypie, _ = ax.pie(group_size, radius=1.3, labels=group_names, colors=colors_grp)
plt.setp( mypie, width=0.3, edgecolor='white')

# Second Ring (Inside)
mypie2, _ = ax.pie(subgroup_size, radius=1.3-0.3, labels=subgroup_names, labeldistance=0.6, colors=colors,rotatelabels = True )
plt.setp( mypie2, width=0.4, edgecolor='white')
plt.margins(0,5)
 

#TOO FANCY
#wedges, texts = ax.pie(group_size, wedgeprops=dict(width=0.5), startangle=-40)
#
#bbox_props = dict(boxstyle="square,pad=0.3", fc="w", ec="k", lw=0.72)
#kw = dict(arrowprops=dict(arrowstyle="-"),
#          bbox=bbox_props, zorder=0, va="center")
#
#for i, p in enumerate(wedges):
#    ang = (p.theta2 - p.theta1)/2. + p.theta1
#    y = np.sin(np.deg2rad(ang))
#    x = np.cos(np.deg2rad(ang))
#    horizontalalignment = {-1: "right", 1: "left"}[int(np.sign(x))]
#    connectionstyle = "angle,angleA=0,angleB={}".format(ang)
#    kw["arrowprops"].update({"connectionstyle": connectionstyle})
#    ax.annotate(group_names[i], xy=(x, y), xytext=(1.35*np.sign(x), 1.4*y),
#                horizontalalignment=horizontalalignment, **kw)
#
#ax.set_title("Antimicrobial Classes: A donut")


plt.title("Antimicrobial class on the outer doughnut and antimicrobials on the inner doughnut (values=number of non-susceptible samples", loc='left')









# show it
plt.show()

#%%


#%%

import pandas as pd
numrows= cursor.execute("SELECT host,count(host) FROM Klebsiellaast GROUP BY(host) ")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
#print(tabulate(rows, headers=['AST_phenotype'], tablefmt='psql'))
df_host = pd.DataFrame(rows)
df_host.columns=['host','count']


import matplotlib.pyplot as plt

# Data to plot
labels =[]
labels= list(df_host["host"])
count=[]
count=list(df_host["count"])
colors = ['gold', 'yellowgreen', 'lightcoral', 'lightskyblue','orange','brown','violet']
#explode = (0.1, 0, 0, 0)  # explode 1st slice
def make_autopct(count):
    def my_autopct(pct):
        total = sum(count)
        val = int(round(pct*total/100.0))
        return '{p:.2f}%  ({v:d})'.format(p=pct,v=val)
    return my_autopct
# Plot
plt.pie(count, radius=1,labels=labels, colors=colors,autopct=make_autopct(count))
plt.title("Host vs number of samples from each host tested for AST", loc='left')
plt.axis('equal')
plt.show()





#%%

#%%

import pandas as pd
numrows= cursor.execute("select AST_phenotypes ,count(`#label`),host from Klebsiellaast group by AMR_genotypes")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
#print(tabulate(rows, headers=['AST_phenotype'], tablefmt='psql'))
df_host = pd.DataFrame(rows)
df_host.columns=['AST_phenotypes','AMR_genotypes'  , 'count',]
total=df_host['count'].sum()
print(total)
df_host=df_host[df_host['count'] >= 160]
df_host=df_host.reset_index()
df_host=df_host.drop('index',axis=1)

#host=[]
#host=list(df_host["count"])
#ast=[]
#ast=list(df_host["AST_phenotypes"])
#ax = df_host.plot.barh(y="count", x="AST_phenotypes")
#plt.title("number of unique AST_phenotypes signatures (>160) total number of unique genotypes were 887(11 in this plot)", loc='left')
#plt.show()





#%%





#%%

import pandas as pd
numrows= cursor.execute("select AST_phenotypes ,count(`#label`) from Klebsiella_ast_us group by AST_phenotypes")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
#print(tabulate(rows, headers=['AST_phenotype'], tablefmt='psql'))
df_host = pd.DataFrame(rows)
df_host.columns=['AST_phenotypes'  , 'count',]
total=df_host['count'].sum()
print(total)
df_host=df_host[df_host['count'] >= 160]
df_host=df_host.reset_index()
df_host=df_host.drop('index',axis=1)

host=[]
host=list(df_host["count"])
ast=[]
ast=list(df_host["AST_phenotypes"])
ax = df_host.plot.barh(y="count", x="AST_phenotypes")
plt.title("number of unique AST_phenotypes signatures (>160) total number of unique genotypes were 887(11 in this plot)", loc='left')
plt.show()





#%%





#%%
import matplotlib.pyplot as plt
import pandas as pd
numrows= cursor.execute("select AST_phenotypes , AMR_genotypes ,count(`#label`) from Klebsiella_ast_us group by AMR_genotypes")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
#print(tabulate(rows, headers=['AST_phenotype'], tablefmt='psql'))
df_host = pd.DataFrame(rows)
df_host.columns=['AST_phenotypes','AMR_genotypes'  , 'count',]
total=df_host['count'].sum()
print(total)
df_host=df_host[df_host['count'] >= 160]
df_host=df_host.reset_index()
df_host=df_host.drop('index',axis=1)
#df_host=df_host.set_index(['count'])
#host=[]
#host=list(df_host["count"])
ast=[]
ast=list(df_host["AST_phenotypes"])
#ax = df_host.plot.barh()
#plt.title("number of unique AST_phenotypes signatures (>160) total number of unique genotypes were 887(11 in this plot)", loc='left')
#plt.show()
df_host=df_host.fillna(value="none", inplace=False)
ind = np.arange(len(df_host))
width = 0.4

fig, ax = plt.subplots()
ax.barh(ind, df_host.AST_phenotypes, width, color='red', label='N')
ax.barh(ind + width, df_host.AMR_genotypes, width, color='green', label='M')
#
ax.set(yticks=ind + width, yticklabels=df_host.count, ylim=[2*width - 1, len(df_host)])
ax.legend()


#%%

#%%

import pandas as pd
numrows= cursor.execute("select lat_lon,count(lat_lon), bioproject_center, count(bioproject_center) , geo_loc_name, count(geo_loc_name), `#label`  from Klebsiella_ast_us group by `#label`")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
print(tabulate(rows, headers=['lat_lon','count(lat_lon)', 'bioproject_center', 'count(bioproject_center)' , 'geo_loc_name', 'count(geo_loc_name)', 'label'], tablefmt='psql'))
df_location = pd.DataFrame(rows)
df_location.columns=['lat_lon','count(lat_lon)', 'bioproject_center', 'count(bioproject_center)' , 'geo_loc_name', 'count(geo_loc_name)', 'label']





#%%






#%%

import pandas as pd
numrows= cursor.execute("select lat_lon, bioproject_center, geo_loc_name, count(`#label`),AST_phenotypes  from Klebsiella_ast_us group by bioproject_center")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
print(tabulate(rows, headers=['lat_lon', 'geo_loc_name','number_of_samples','AST_phenotypes'], tablefmt='psql'))
df_location = pd.DataFrame(rows)
df_location.columns=['lat_lon', 'bioproject_center' , 'geo_loc_name', 'number_of_samples','AST_phenotypes']





#%%


#%%

import pandas as pd
numrows= cursor.execute("select AST_phenotypes,count(AST_phenotypes), bioproject_center from Klebsiella_ast_us group by bioproject_center")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
print(tabulate(rows, headers=[ 'AST_phenotypes','number_of_samples','bioproject_center'], tablefmt='psql'))
df_location = pd.DataFrame(rows)
df_location.columns=['AST_phenotypes','number_of_samples','bioproject_center']





#%%

#%%

import pandas as pd
numrows= cursor.execute("select AST_phenotypes,count(AST_phenotypes), bioproject_center from Klebsiella_ast_us group by AST_phenotypes")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
print(tabulate(rows, headers=[ 'AST_phenotypes','number_of_samples','bioproject_center'], tablefmt='psql'))
df_location = pd.DataFrame(rows)
df_location.columns=['AST_phenotypes','number_of_samples','bioproject_center']





#%%


#%%

import pandas as pd
numrows= cursor.execute("select AST_phenotypes,count(AST_phenotypes), geo_loc_name from Klebsiella_ast_us_us group by geo_loc_name")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
print(tabulate(rows, headers=[ 'AST_phenotypes','number_of_samples','geo_loc_name'], tablefmt='psql'))
df_location = pd.DataFrame(rows)
df_location.columns=['AST_phenotypes','number_of_samples','geo_loc_name']

ax = df_location.plot.barh(y="number_of_samples", x="geo_loc_name")
plt.title("number of samples from each geolocation name", loc='left')
plt.figtext(0.15,0.02,"SQL QUERY:select AST_phenotypes,count(AST_phenotypes), geo_loc_name from Klebsiellaast group by geo_loc_name")

plt.show()



#%%
#THis has the same values as group by AST_phenotypes
#%%

import pandas as pd
numrows= cursor.execute("select bioproject_center, AST_phenotypes,count(AST_phenotypes) from Klebsiella_ast_us group by bioproject_center")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
print(tabulate(rows, headers=[ 'bioproject_center','number_of_samples'], tablefmt='psql'))
df_location = pd.DataFrame(rows)
df_location.columns=['bioproject_center','AST_phenotypes','number_of_samples']

ax = df_location.plot.barh(y="number_of_samples", x="bioproject_center")
plt.title("Number of samples from bioproject_centers tested for AST", loc='left')

plt.figtext(0.15,0.02,"SQL QUERY:select bioproject_center, AST_phenotypes,count(AST_phenotypes), from Klebsiellaast group by bioproject_center")
plt.show()


#%%
#this has same values as group by bioproject_center
#%%

import pandas as pd
numrows= cursor.execute("select bioproject_center, count(bioproject_center) from Klebsiella_ast_us group by bioproject_center")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
print(tabulate(rows, headers=[ 'bioproject_center','number_of_samples'], tablefmt='psql'))
df_bioproj = pd.DataFrame(rows)
df_bioproj.columns=['bioproject_center','number_of_samples']

ax = df_bioproj.plot.barh(y="number_of_samples", x="bioproject_center")
plt.title("number of samples from each bioproject center", loc='left')
plt.figtext(0.15,0.02,"select bioproject_center, count(bioproject_center) from Klebsiellaast group by bioproject_center")

plt.show()




#%%



#%%

import pandas as pd
numrows= cursor.execute("select bioproject_center,geo_loc_name,lat_lon, AST_phenotypes, count(AST_phenotypes) from Klebsiella_ast_us group by AST_phenotypes")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
print(tabulate(rows, headers=[ 'bioproject_center','number_of_samples'], tablefmt='psql'))
df_location_all = pd.DataFrame(rows)
df_location_all.columns=['bioproject_center','geo_loc_name','lat_lon','AST_phenotypes','number_of_samples']

#ax = df_location.plot.barh(y="number_of_samples", x="bioproject_center")
#plt.title("number of samples from each bioproject center", loc='left')
#plt.show()




#%%

#%%

import pandas as pd
numrows= cursor.execute("select AST_phenotypes, count(AST_phenotypes) from Klebsiella_ast_us group by AST_phenotypes")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
print(tabulate(rows, headers=[ 'bioproject_center','number_of_samples'], tablefmt='psql'))
df_phenotype = pd.DataFrame(rows)
df_phenotype.columns=['AST_phenotypes','number_of_samples']

#ax = df_location.plot.barh(y="number_of_samples", x="bioproject_center")
#plt.title("number of samples from each bioproject center", loc='left')
#plt.show()




#%%

#origins of each ast_phenotype grouped by ast_phenotype
#%%

import pandas as pd
phenotype=[]
appended_phenotype=[]
phenotype=list(df_phenotype["AST_phenotypes"])
for x in phenotype:
    numrows= cursor.execute("select geo_loc_name,bioproject_center, count(AST_phenotypes) from Klebsiella_ast_us where AST_phenotypes ='" + x + "'")
    print("Selected %s rows" %numrows)
    print("Selected %s rows " %cursor.rowcount)
    rows =cursor.fetchall()#fetch all rows at once
    print(tabulate(rows, headers=[ 'bioproject_center','number_of_samples'], tablefmt='psql'))
    appended_phenotype.append(rows) #appending tuple result into a list
#converting list of tuples into a dataframe
df_location= pd.DataFrame(appended_phenotype)
df_location.columns=["pheno"]
df1=pd.DataFrame(df_location)
df1[['geo_loc_name','bioproject_center','number_of_samples']] = pd.DataFrame(df1.pheno.values.tolist(), index= df1.index)
df_grouped_phenotype = pd.DataFrame(df1['pheno'].values.tolist(), columns=['geo_loc_name','bioproject_center','number_of_samples'])


#%%

#origins of each ast_phenotype grouped by geo_loc and bioproject center
#%%

import pandas as pd
bioproj=[]
appended_bioproj=[]
bioproj=list(df_bioproj["bioproject_center"])
for x in bioproj:
    try:
        numrows= cursor.execute("select geo_loc_name,bioproject_center, count(AST_phenotypes),AST_phenotypes from Klebsiella_ast_us where bioproject_center = '"  + x + " ' group by geo_loc_name")
        print("Selected %s rows" %numrows)
        print("Selected %s rows " %cursor.rowcount)
        rows =cursor.fetchall()#fetch all rows at once
        print(tabulate(rows, headers=[ 'a','b','bioproject_center','number_of_samples'], tablefmt='psql'))
        appended_bioproj.append(rows) #appending tuple result into a list]
    except:
        print("women's hospital was removed")
        continue
    
#converting list of tuples into a dataframe
df_location= pd.DataFrame(appended_bioproj)
df_location.columns=["pheno"]
#df1=pd.DataFrame(df_location)
#df1[['geo_loc_name','bioproject_center','number_of_samples','AST_phenotype']] = pd.DataFrame(df1.pheno.values.tolist(), index= df1.index)


#df_curr = pd.DataFrame(rows)
#df_curr.columns=['geo_loc_name','bioproject_center','number_of_samples','AST_phenotype']



#%%


#%%

import pandas as pd

numrows= cursor.execute("select geo_loc_name,bioproject_center, count(AST_phenotypes),AST_phenotypes from Klebsiella_ast_us_us where bioproject_center = 'Broad Institute' group by geo_loc_name")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
print(tabulate(rows, headers=[ 'a','b','bioproject_center','number_of_samples'], tablefmt='psql'))


df_curr = pd.DataFrame(rows)
df_curr.columns=['geo_loc_name','bioproject_center','number_of_samples','AST_phenotypes']

ax = df_curr.plot.barh(y="number_of_samples", x="AST_phenotypes")
plt.title("number of samples from each bioproject center/geolocation", loc='left')
plt.figtext(0.15,0.02,"SQL Quer:select geo_loc_name,bioproject_center, count(AST_phenotypes),AST_phenotypes from Klebsiellaast where bioproject_center = 'Broad Institute' group by geo_loc_name")

plt.show()


#%%

#%%

import pandas as pd

numrows= cursor.execute("select count(`#label`), host, isolation_source from Klebsiella_ast_us group by isolation_source;")
print("Selected %s rows" %numrows)
print("Selected %s rows " %cursor.rowcount)
rows =cursor.fetchall()#fetch all rows at once
print(tabulate(rows, headers=[ 'Number_of_samples','Host','Isolation_source'], tablefmt='psql'))


df_curr = pd.DataFrame(rows)
df_curr.columns=['Number_of_samples','Host','Isolation_source']




#%%